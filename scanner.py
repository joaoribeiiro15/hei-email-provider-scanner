#!/usr/bin/env python3
"""
Email Provider Scanner
Identifies the email provider of each institution by querying MX and TXT (SPF)
DNS records via the Cloudflare DNS over HTTPS API (RFC 8484).

No API key required. No third-party dependencies.
API: https://developers.cloudflare.com/1.1.1.1/encryption/dns-over-https/make-api-requests/dns-json/
"""

import csv
import json
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DOH_ENDPOINT = "https://cloudflare-dns.com/dns-query"
REQUEST_DELAY = 0.5          # seconds between requests (polite to the public API)
RESULTS_DIR   = Path("results")
SOURCE_DIR    = Path("source")

# ---------------------------------------------------------------------------
# Email provider fingerprints
# Matched against MX hostnames (and TXT/SPF records as secondary evidence).
# Order matters: more specific patterns first.
# ---------------------------------------------------------------------------

MX_FINGERPRINTS = [
    # Microsoft 365 / Exchange Online
    ("Microsoft 365",       ["mail.protection.outlook.com"]),
    # Google Workspace (formerly G Suite)
    ("Google Workspace",    ["aspmx.l.google.com", "googlemail.com", "smtp.google.com"]),
    # Proofpoint (security gateway, often fronts M365 or GW)
    ("Proofpoint",          ["pphosted.com", "proofpoint.com"]),
    # Mimecast (security gateway)
    ("Mimecast",            ["mimecast.com"]),
    # Barracuda (security gateway)
    ("Barracuda",           ["barracudanetworks.com", "cudamail.com"]),
    # Cisco / IronPort
    ("Cisco / IronPort",    ["iphmx.com", "cisco.com"]),
    # Sophos / UTM
    ("Sophos",              ["sophos.com", "himgs.com"]),
    # Forcepoint
    ("Forcepoint",          ["forcepoint.com", "mailcontrol.com"]),
    # SpamExperts / Leaseweb
    ("SpamExperts",         ["spamexperts.com", "antispamcloud.com"]),
    # Zoho Mail
    ("Zoho Mail",           ["zoho.com", "zohomail.com"]),
    # Amazon SES / WorkMail
    ("Amazon WorkMail",     ["awsapps.com", "mail.eu-west", "inbound-smtp"]),
    # Yahoo
    ("Yahoo Mail",          ["yahoodns.net", "yahoo.com"]),
    # Apple iCloud
    ("Apple iCloud",        ["icloud.com", "apple.com"]),
    # FastMail
    ("FastMail",            ["fastmail.com", "fastmail.fm"]),
    # Hetzner
    ("Hetzner",             ["your-server.de", "hetzner.com"]),
    # Tuta / Tutanota
    ("Tutanota",            ["tutanota.de", "tutanota.com", "tuta.io"]),
    # ProtonMail
    ("ProtonMail",          ["protonmail.ch", "proton.me"]),
    # Norwegian university / government self-hosted patterns
    ("UNINETT / Sikt",      ["uninett.no", "sikt.no"]),
    # Generic self-hosted indicators (catch-all: domain matches own MX)
]

SPF_FINGERPRINTS = [
    ("Microsoft 365",    ["include:spf.protection.outlook.com", "include:protection.outlook.com"]),
    ("Google Workspace", ["include:_spf.google.com", "include:spf.google.com"]),
    ("Proofpoint",       ["include:pphosted.com"]),
    ("Mimecast",         ["include:spf.mimecast.com"]),
    ("Zoho Mail",        ["include:zoho.com", "include:transmail.net"]),
    ("Amazon SES",       ["include:amazonses.com"]),
    ("Tutanota",         ["include:tutanota.de"]),
    ("ProtonMail",       ["include:_spf.protonmail.ch"]),
    ("UNINETT / Sikt",   ["include:uninett.no", "include:sikt.no"]),
]

# ---------------------------------------------------------------------------
# Cloudflare DoH helper
# ---------------------------------------------------------------------------

def _doh_query(name: str, rtype: str) -> list[dict]:
    """
    Query Cloudflare DNS over HTTPS (JSON format).
    Returns the Answer list, or [] on any error.
    API: GET https://cloudflare-dns.com/dns-query?name=<name>&type=<type>
    """
    url = f"{DOH_ENDPOINT}?name={name}&type={rtype}"
    req = Request(
        url,
        headers={
            "Accept": "application/dns-json",
            "User-Agent": "mx-scanner/1.0",
        },
    )
    try:
        with urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
            return data.get("Answer", [])
    except (HTTPError, URLError, json.JSONDecodeError):
        return []


def get_mx_records(domain: str) -> list[tuple[int, str]]:
    """Return sorted list of (priority, hostname) MX records for domain."""
    answers = _doh_query(domain, "MX")
    records = []
    for ans in answers:
        if ans.get("type") == 15:       # DNS type 15 = MX
            raw = ans.get("data", "")   # e.g. "10 aspmx.l.google.com."
            parts = raw.strip().split(None, 1)
            if len(parts) == 2:
                try:
                    priority = int(parts[0])
                    hostname = parts[1].rstrip(".")
                    records.append((priority, hostname.lower()))
                except ValueError:
                    pass
    records.sort(key=lambda x: x[0])
    return records


def get_txt_records(domain: str) -> list[str]:
    """Return TXT record strings for domain (used to read SPF)."""
    answers = _doh_query(domain, "TXT")
    result = []
    for ans in answers:
        if ans.get("type") == 16:       # DNS type 16 = TXT
            raw = ans.get("data", "").strip('"')
            result.append(raw.lower())
    return result

# ---------------------------------------------------------------------------
# Provider detection
# ---------------------------------------------------------------------------

def detect_provider_from_mx(mx_records: list[tuple[int, str]]) -> tuple[str, str]:
    """
    Match MX hostnames against known fingerprints.
    Returns (provider_name, matched_mx_hostname).
    """
    for priority, hostname in mx_records:
        for provider, patterns in MX_FINGERPRINTS:
            for pattern in patterns:
                if pattern in hostname:
                    return provider, hostname
    return "", ""


def detect_provider_from_spf(txt_records: list[str]) -> str:
    """
    Try to identify provider from SPF TXT records as secondary evidence.
    Returns provider_name or "".
    """
    spf_records = [t for t in txt_records if t.startswith("v=spf1")]
    for spf in spf_records:
        for provider, patterns in SPF_FINGERPRINTS:
            for pattern in patterns:
                if pattern in spf:
                    return provider
    return ""


def classify_provider(domain: str, mx_records: list[tuple[int, str]], txt_records: list[str]) -> dict:
    """
    Determine the email provider with confidence and evidence.
    Returns a dict with all relevant fields.
    """
    if not mx_records:
        return {
            "provider":     "No MX record",
            "confidence":   "N/A",
            "evidence":     "No MX records returned by DNS",
            "mx_primary":   "",
            "mx_all":       "",
            "spf_record":   "",
            "note":         "Domain may not have email configured.",
        }

    mx_all = "; ".join(f"{p} {h}" for p, h in mx_records)
    primary_mx = mx_records[0][1] if mx_records else ""

    spf_list = [t for t in txt_records if t.startswith("v=spf1")]
    spf_record = spf_list[0] if spf_list else ""

    mx_provider, mx_match = detect_provider_from_mx(mx_records)
    spf_provider = detect_provider_from_spf(txt_records)

    # Self-hosted heuristic: primary MX is a subdomain of the institution's own domain
    base_domain = ".".join(domain.split(".")[-2:])
    is_self_hosted = base_domain in primary_mx and not mx_provider

    if mx_provider:
        if spf_provider and spf_provider != mx_provider:
            # MX provider differs from SPF: likely a security gateway in front
            note = f"Security gateway ({mx_provider}) fronting {spf_provider}. Actual mailboxes probably at {spf_provider}."
            provider = f"{mx_provider} (gateway) / {spf_provider}"
            confidence = "High"
            evidence = f"MX matches '{mx_match}'; SPF includes '{spf_provider}'"
        else:
            provider = mx_provider
            confidence = "High"
            evidence = f"MX hostname '{mx_match}' matches known fingerprint"
            note = ""
    elif spf_provider:
        provider = spf_provider
        confidence = "Medium"
        evidence = f"Detected via SPF record only (MX did not match known patterns)"
        note = "MX pattern unknown; provider inferred from SPF."
    elif is_self_hosted:
        provider = "Self-hosted"
        confidence = "Medium"
        evidence = f"Primary MX '{primary_mx}' is a subdomain of the institution's own domain"
        note = "Institution appears to run its own mail server."
    else:
        provider = "Unknown"
        confidence = "Low"
        evidence = f"Primary MX '{primary_mx}' did not match any known provider"
        note = "Manual investigation recommended."

    return {
        "provider":   provider,
        "confidence": confidence,
        "evidence":   evidence,
        "mx_primary": primary_mx,
        "mx_all":     mx_all,
        "spf_record": spf_record,
        "note":       note if "note" in dir() else "",
    }

# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------

def write_xlsx(rows: list[dict], fieldnames: list[str], path: Path) -> None:
    from collections import Counter

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    CONFIDENCE_FILL = {
        "High":   PatternFill("solid", fgColor="C6EFCE"),
        "Medium": PatternFill("solid", fgColor="FFEB9C"),
        "Low":    PatternFill("solid", fgColor="FFC7CE"),
        "N/A":    PatternFill("solid", fgColor="EDEDED"),
    }
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(vertical="center", wrap_text=True)
    title_font = Font(bold=True, size=13, color="1F3864")
    sub_font   = Font(bold=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Data ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Data"

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = center
    ws.row_dimensions[1].height = 22

    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])
        row_idx = ws.max_row
        fill = CONFIDENCE_FILL.get(row.get("confidence", ""))
        for cell in ws[row_idx]:
            cell.alignment = left
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 10

    # Provider distribution table
    provider_counts = Counter(r["provider"] for r in rows)
    sorted_providers = sorted(provider_counts.items(), key=lambda x: -x[1])
    n_prov = len(sorted_providers)

    ws2["A1"] = "Provider Distribution"
    ws2["A1"].font = title_font
    ws2["A2"] = "Provider"
    ws2["B2"] = "Count"
    ws2["A2"].font = sub_font
    ws2["B2"].font = sub_font

    for i, (prov, cnt) in enumerate(sorted_providers, start=3):
        ws2.cell(row=i, column=1, value=prov)
        ws2.cell(row=i, column=2, value=cnt)

    chart1 = BarChart()
    chart1.type = "bar"
    chart1.title = "Email Provider Distribution"
    chart1.y_axis.title = "Provider"
    chart1.x_axis.title = "Institutions"
    chart1.width = 22
    chart1.height = max(10, min(n_prov * 0.9, 20))
    data1 = Reference(ws2, min_col=2, min_row=2, max_row=2 + n_prov)
    cats1 = Reference(ws2, min_col=1, min_row=3, max_row=2 + n_prov)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws2.add_chart(chart1, "D2")

    # Anchor chart2 below chart1 based on its actual height (1 cm ≈ 1.89 rows at default row height)
    chart2_anchor_row = 2 + int(chart1.height / 0.529) + 3

    # Confidence distribution table
    conf_row = n_prov + 4
    conf_counts = Counter(r["confidence"] for r in rows)
    conf_order = ["High", "Medium", "Low", "N/A"]

    ws2.cell(row=conf_row, column=1, value="Confidence Distribution").font = title_font
    ws2.cell(row=conf_row + 1, column=1, value="Confidence").font = sub_font
    ws2.cell(row=conf_row + 1, column=2, value="Count").font = sub_font

    for i, conf in enumerate(conf_order, start=conf_row + 2):
        ws2.cell(row=i, column=1, value=conf)
        ws2.cell(row=i, column=2, value=conf_counts.get(conf, 0))

    chart2 = PieChart()
    chart2.title = "Confidence Level Distribution"
    chart2.width = 16
    chart2.height = 12
    data2 = Reference(ws2, min_col=2, min_row=conf_row + 1, max_row=conf_row + 1 + len(conf_order))
    cats2 = Reference(ws2, min_col=1, min_row=conf_row + 2, max_row=conf_row + 1 + len(conf_order))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    ws2.add_chart(chart2, f"D{chart2_anchor_row}")

    wb.save(path)


# ---------------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------------

def find_source_csvs() -> list[Path]:
    files = sorted(SOURCE_DIR.glob("*.csv"))
    if not files:
        print(f"[WARN] No CSV files found in '{SOURCE_DIR}/'.")
    return files


def read_institutions(csv_path: Path) -> list[dict]:
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def extract_domain(url_value: str) -> str:
    """
    Strip scheme, path, and leading subdomains, returning the registrable
    domain for DNS MX/TXT queries.

    Examples:
      www.uia.no     -> uia.no
      en.uit.no      -> uit.no
      www.ntnu.edu   -> ntnu.edu
    """
    domain = url_value.strip()
    domain = domain.replace("https://", "").replace("http://", "")
    domain = domain.split("/")[0].lower()           # remove path

    # Norwegian ccTLD: keep last two labels (second-level.no)
    # Generic TLDs (.com, .org, .edu, .net): keep last two labels
    # Two-part ccTLD second-levels (.co.uk, .ac.uk, etc.) are rare in .no — not needed
    parts = domain.split(".")
    if len(parts) > 2:
        domain = ".".join(parts[-2:])
    return domain

# ---------------------------------------------------------------------------
# Display helpers
# ---------------------------------------------------------------------------

RESET  = "\033[0m"
BOLD   = "\033[1m"
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
DIM    = "\033[2m"


def confidence_color(conf: str) -> str:
    return {
        "High":   GREEN,
        "Medium": YELLOW,
        "Low":    RED,
        "N/A":    DIM,
    }.get(conf, RESET)


def print_header():
    print()
    print(f"{BOLD}{'='*70}{RESET}")
    print(f"{BOLD}  Email Provider Scanner  —  Cloudflare DoH API{RESET}")
    print(f"  Run at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{BOLD}{'='*70}{RESET}")
    print()


def print_result(idx: int, total: int, name: str, url: str, domain: str, info: dict):
    col = confidence_color(info["confidence"])
    print(f"{BOLD}[{idx}/{total}] {name}{RESET}")
    print(f"  URL            : {url}")
    print(f"  Query domain   : {domain}")
    print(f"  Provider       : {CYAN}{BOLD}{info['provider']}{RESET}")
    print(f"  Confidence     : {col}{info['confidence']}{RESET}")
    print(f"  Evidence       : {info['evidence']}")
    if info.get("note"):
        print(f"  Note           : {DIM}{info['note']}{RESET}")
    if info["mx_primary"]:
        print(f"  Primary MX     : {info['mx_primary']}")
    if info["mx_all"] and "; " in info["mx_all"]:
        print(f"  All MX         : {DIM}{info['mx_all']}{RESET}")
    if info["spf_record"]:
        print(f"  SPF record     : {DIM}{info['spf_record'][:100]}{'...' if len(info['spf_record']) > 100 else ''}{RESET}")
    print(f"  {DIM}{'-'*60}{RESET}")


# ---------------------------------------------------------------------------
# Result row builder
# ---------------------------------------------------------------------------

def build_result_row(institution: dict, domain: str, info: dict) -> dict:
    return {
        "ID":           institution.get("ID", ""),
        "Name":         institution.get("Name", ""),
        "Category":     institution.get("Category", ""),
        "NUTS2":        institution.get("NUTS2", ""),
        "NUTS2_Label":  institution.get("NUTS2_Label", ""),
        "url":          institution.get("url", ""),
        "query_domain": domain,
        "provider":     info["provider"],
        "confidence":   info["confidence"],
        "evidence":     info["evidence"],
        "note":         info.get("note", ""),
        "mx_primary":   info["mx_primary"],
        "mx_all":       info["mx_all"],
        "spf_record":   info["spf_record"],
    }

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print_header()

    RESULTS_DIR.mkdir(exist_ok=True)

    csv_files = find_source_csvs()
    if not csv_files:
        sys.exit(1)

    all_rows = []

    for csv_path in csv_files:
        print(f"{BOLD}Source file: {csv_path}{RESET}\n")
        institutions = read_institutions(csv_path)
        total = len(institutions)
        print(f"  Found {total} institutions.\n")

        for idx, inst in enumerate(institutions, 1):
            url_raw = inst.get("url", "").strip()
            name    = inst.get("Name", url_raw)

            if not url_raw:
                print(f"[{idx}/{total}] {name} — SKIPPED (no URL)\n")
                continue

            domain = extract_domain(url_raw)

            mx_records  = get_mx_records(domain)
            time.sleep(REQUEST_DELAY)
            txt_records = get_txt_records(domain)
            time.sleep(REQUEST_DELAY)

            info = classify_provider(domain, mx_records, txt_records)
            print_result(idx, total, name, url_raw, domain, info)

            all_rows.append(build_result_row(inst, domain, info))

    # Write CSV and XLSX
    if all_rows:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fieldnames = list(all_rows[0].keys())

        out_path = RESULTS_DIR / f"email_providers_{ts}.csv"
        with open(out_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_rows)

        xlsx_path = RESULTS_DIR / f"email_providers_{ts}.xlsx"
        write_xlsx(all_rows, fieldnames, xlsx_path)

        print(f"\n{GREEN}{BOLD}Results saved to: {out_path}{RESET}")
        print(f"{GREEN}{BOLD}Results saved to: {xlsx_path}{RESET}")
        print(f"Total institutions analysed: {len(all_rows)}\n")

        # Provider summary
        from collections import Counter
        provider_counts = Counter(r["provider"] for r in all_rows)
        print(f"{BOLD}Provider distribution:{RESET}")
        for provider, count in sorted(provider_counts.items(), key=lambda x: -x[1]):
            bar = "█" * count
            print(f"  {CYAN}{provider:<40}{RESET}  {count:>2}  {DIM}{bar}{RESET}")

    print(f"\n{BOLD}Done.{RESET}\n")


if __name__ == "__main__":
    main()
